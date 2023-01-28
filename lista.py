import openpyxl

# Cria uma nova planilha
wb = openpyxl.Workbook()
sheet = wb.active

# Listas de nome, código e situação
nome = ["João", "Maria", "José"]
codigo = [1, 2, 3]
situacao = ["Ativo", "Inativo", "Ativo"]

# Adiciona os cabeçalhos às colunas
sheet.cell(row=1, column=1).value = "Nome"
sheet.cell(row=1, column=2).value = "Código"
sheet.cell(row=1, column=3).value = "Situação"

# Adiciona os dados às colunas
for i in range(len(nome)):
    sheet.cell(row=i+2, column=1).value = nome[i]
    sheet.cell(row=i+2, column=2).value = codigo[i]
    sheet.cell(row=i+2, column=3).value = situacao[i]

# Salva a planilha como um arquivo Excel
wb.save("dados.xlsx")
